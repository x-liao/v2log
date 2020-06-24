#!/usr/bin/env python3
import os
import sys
import json
import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Alignment


path = os.path.split(os.path.realpath(__file__))[0]
log_path = os.path.split(os.path.realpath(__file__))[0] + '/logs'


def exte(a,*endstring):
	array = map(a.endswith,endstring)
	if True in array:
		return True
	else:
		return False

def save_count_log(filename,data):
	with open(filename,'w') as f:
		f.write(data)

def get_files():
	type = '.json'
	data = {}
	for dirpath,dirnames,filenames in os.walk(log_path):
		host = (os.path.split(dirpath))[1]
		if host == 'logs':
			continue
		files = []
		for name in filenames:
			if exte(name,type):
				json_file = os.path.join(dirpath, name)
				files.append(json_file)
		data[host] = files
	return data

def save_excel(name,domains):
	wb = Workbook()
	dest_filename = name + '.xlsx'
	ws = wb.active
	ws.title = name

	ws.cell(column=1,row=1,value='域名')
	ws.cell(column=1,row=2,value='协议')
	row = 2
	date_value = 2
	date_dic = {}
	alignment = Alignment(horizontal='center', vertical='center')
	for domain in domains.keys():
		domain_value = domains[domain]
		row += 1
		# 写域名
		ws.cell(column=1,row=row,value=domain)

		for protocol in domain_value.keys():
			# TCP数据
			if protocol == 'tcp':
				count_data = domain_value[protocol]
				for date in count_data.keys():
					date_column = date_dic.setdefault(date, 0)
					if not date_column:
						date_dic[date] = date_value
						ws.cell(column=date_dic[date],row=1,value=date).alignment = alignment
						ws.merge_cells(start_column=date_dic[date], end_column=date_dic[date] + 3,start_row=1, end_row=1)
						

						ws.cell(column=date_dic[date],row=2,value='tcp').alignment = alignment
						ws.merge_cells(start_column=date_dic[date], end_column=date_dic[date] + 1,start_row=2, end_row=2)

						ws.cell(column=date_dic[date] + 2,row=2,value='udp').alignment = alignment
						ws.merge_cells(start_column=date_dic[date] + 2, end_column=date_dic[date] + 3,start_row=2, end_row=2)
						date_value += 4

					ws.cell(column=date_dic[date],row=row,value=count_data[date][0])
					ws.cell(column=date_dic[date] + 1,row=row,value=count_data[date][1])

			# UDP数据
			if protocol == 'udp':
				count_data = domain_value[protocol]
				for date in count_data.keys():
					date_column = date_dic.setdefault(date, 0)

					if not date_column:
						date_dic[date] = date_value

						ws.cell(column=date_dic[date],row=1,value=date).alignment = alignment
						ws.merge_cells(start_column=date_dic[date], end_column=date_dic[date] + 3,start_row=1, end_row=1)

						ws.cell(column=date_dic[date],row=2,value='tcp').alignment = alignment
						ws.merge_cells(start_column=date_dic[date], end_column=date_dic[date] + 1,start_row=2, end_row=2)

						ws.cell(column=date_dic[date] + 2,row=2,value='udp').alignment = alignment
						ws.merge_cells(start_column=date_dic[date] + 2, end_column=date_dic[date] + 3,start_row=2, end_row=2)
						date_value += 4

					ws.cell(column=date_dic[date] + 2,row=row,value=count_data[date][0])
					ws.cell(column=date_dic[date] + 3,row=row,value=count_data[date][1])



	wb.save(filename=dest_filename)

if __name__ == '__main__':
	files = get_files()
	print('一共%d个文件'%len(files))

	for key in files.keys():
		name = key
		count_data = defaultdict(dict)
		for file in files[name]:
			with open(file, 'r') as f:
				data = f.read()
			data = json.loads(data)
			create_date = data['create_date'].split('-')
			create_date = '-'.join(create_date[:3])

			for item in data['data']:
				domain = item['domain']
				count_item = count_data[domain]
				count_item[item['protocol']] = count_item.setdefault(item['protocol'], {})
				count_item[item['protocol']][create_date] = count_item[item['protocol']].setdefault(create_date, {})
				protocol_data = count_item[item['protocol']][create_date]
				try:
					count_item[item['protocol']][create_date] = [item['port'],protocol_data[1] + item['count']]
				except:
					count_item[item['protocol']][create_date] = [item['port'],item['count']]
				
				count_data[domain] = count_item

		with open('%s/data/%s.json'%(path,name),'w') as f:
			data = json.dumps(count_data)
			f.write(data)

		save_excel(name,count_data)
